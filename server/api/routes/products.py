import io
import csv
import json
from datetime import datetime
from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, asc, func, or_
from loguru import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.database import get_db
from core.models import Product, PriceHistory
from core.schemas import ProductResponse, PriceHistoryResponse
from core.serializers import serialize_product, serialize_price_history

router = APIRouter(prefix="/products", tags=["Product Catalog"])


@router.get("", summary="Tìm kiếm & lọc danh sách sản phẩm")
async def list_products(
    q: Optional[str] = Query(None, description="Từ khóa tìm kiếm tên/thương hiệu"),
    platform: Optional[str] = Query(None, description="Sàn TMĐT: lazada"),
    min_price: Optional[float] = Query(None, description="Giá tối thiểu"),
    max_price: Optional[float] = Query(None, description="Giá tối đa"),
    has_ai: Optional[bool] = Query(None, description="Lọc theo trạng thái đã có phân tích AI"),
    sort_by: str = Query("created_desc", description="Sắp xếp: price_asc, price_desc, sold_desc, rating_desc, created_desc"),
    page: int = Query(1, ge=1, description="Số trang"),
    page_size: int = Query(20, ge=1, le=100, description="Số sản phẩm mỗi trang"),
    db: AsyncSession = Depends(get_db)
):
    query = select(Product)

    # 1. Bộ lọc từ khóa
    if q and q.strip():
        search_term = f"%{q.strip()}%"
        query = query.where(
            or_(
                Product.name.ilike(search_term),
                Product.brand.ilike(search_term),
                Product.category.ilike(search_term)
            )
        )

    # 2. Bộ lọc sàn
    if platform and platform.lower() != "all":
        query = query.where(Product.platform == platform.lower())

    # 3. Bộ lọc giá
    if min_price is not None:
        query = query.where(Product.current_price >= min_price)
    if max_price is not None:
        query = query.where(Product.current_price <= max_price)

    # 4. Bộ lọc trạng thái phân tích AI
    if has_ai is not None:
        if has_ai:
            query = query.where(Product.ai_analysis.isnot(None))
        else:
            query = query.where(Product.ai_analysis.is_(None))

    # Đếm tổng số bản ghi
    count_stmt = select(func.count()).select_from(query.subquery())
    total_count = (await db.execute(count_stmt)).scalar() or 0

    # 5. Sắp xếp
    if sort_by == "price_asc":
        query = query.order_by(asc(Product.current_price))
    elif sort_by == "price_desc":
        query = query.order_by(desc(Product.current_price))
    elif sort_by == "sold_desc":
        query = query.order_by(desc(Product.historical_sold))
    elif sort_by == "rating_desc":
        query = query.order_by(desc(Product.rating_star))
    else:
        query = query.order_by(desc(Product.id))

    # 6. Phân trang
    offset = (page - 1) * page_size
    query = query.offset(offset).limit(page_size)

    result = await db.execute(query)
    products = result.scalars().all()

    return {
        "items": [serialize_product(p) for p in products],
        "total": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": (total_count + page_size - 1) // page_size if total_count > 0 else 1
    }


@router.get("/export/csv", summary="Xuất danh sách sản phẩm ra tệp CSV (UTF-8 BOM chuẩn tiếng Việt)")
async def export_products_csv(
    q: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Xuất toàn bộ danh sách sản phẩm kèm dữ liệu phân tích AI ra CSV"""
    query = select(Product).order_by(desc(Product.id))
    if q and q.strip():
        search_term = f"%{q.strip()}%"
        query = query.where(or_(Product.name.ilike(search_term), Product.brand.ilike(search_term)))
    if platform and platform.lower() != "all":
        query = query.where(Product.platform == platform.lower())

    result = await db.execute(query)
    products = result.scalars().all()

    output = io.StringIO()
    # Ghi UTF-8 BOM để Excel không lỗi font tiếng Việt
    output.write('\ufeff')
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)

    # Tiêu đề cột
    writer.writerow([
        "ID",
        "Tên Sản Phẩm Gốc",
        "Tên Chuẩn Hóa (AI)",
        "Sàn TMĐT",
        "Giá Hiện Tại (VNĐ)",
        "Giá Gốc (VNĐ)",
        "Giảm Giá (%)",
        "Đã Bán",
        "Đánh Giá Sao",
        "Số Lượt Đánh Giá",
        "Gian Hàng",
        "Tóm Tắt Chất Lượng (AI)",
        "Ưu Điểm Nổi Bật (AI)",
        "Nhược Điểm / Lưu Ý (AI)",
        "Điểm Cảm Xúc (Thang 10)",
        "Giá Gợi Ý Tối Ưu (VNĐ)",
        "Lời Khuyên Quyết Định (AI)",
        "Link Sản Phẩm Lazada"
    ])

    for p in products:
        ai = p.ai_analysis or {}
        pros_str = " | ".join(ai.get("pros", [])) if isinstance(ai, dict) else ""
        cons_str = " | ".join(ai.get("cons", [])) if isinstance(ai, dict) else ""
        norm_name = ai.get("normalized_name", "") if isinstance(ai, dict) else ""
        quality = ai.get("quality_summary", "") if isinstance(ai, dict) else ""
        sentiment = ai.get("sentiment_score", "") if isinstance(ai, dict) else ""
        p_opt = ai.get("recommended_price_optimal", "") if isinstance(ai, dict) else ""
        verdict = ai.get("buying_verdict", "") if isinstance(ai, dict) else ""

        writer.writerow([
            p.id,
            p.name,
            norm_name,
            p.platform.upper(),
            f"{p.current_price:,.0f}",
            f"{p.original_price:,.0f}" if p.original_price else "",
            f"{p.discount_percentage:.1f}%" if p.discount_percentage else "0%",
            p.historical_sold or 0,
            p.rating_star or 0,
            p.rating_count or 0,
            p.shop_name or "",
            quality,
            pros_str,
            cons_str,
            sentiment,
            f"{p_opt:,.0f}" if p_opt else "",
            verdict,
            p.url
        ])

    csv_data = output.getvalue().encode('utf-8-sig')
    filename = f"lazada_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        io.BytesIO(csv_data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/excel", summary="Xuất danh sách sản phẩm ra tệp Excel (.xlsx) chuyên nghiệp")
async def export_products_excel(
    q: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Xuất danh sách sản phẩm và phân tích AI ra định dạng bảng Excel (.xlsx) được format đẹp mắt"""
    query = select(Product).order_by(desc(Product.id))
    if q and q.strip():
        search_term = f"%{q.strip()}%"
        query = query.where(or_(Product.name.ilike(search_term), Product.brand.ilike(search_term)))
    if platform and platform.lower() != "all":
        query = query.where(Product.platform == platform.lower())

    result = await db.execute(query)
    products = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lazada Products Report"

    # Header styling
    headers = [
        "ID", "Tên Sản Phẩm Gốc", "Tên Chuẩn Hóa (AI)", "Sàn", "Giá Hiện Tại (đ)", 
        "Giá Gốc (đ)", "Giảm Giá", "Đã Bán", "Đánh Giá", "Gian Hàng", 
        "Tóm Tắt Chất Lượng (AI)", "Ưu Điểm (AI)", "Nhược Điểm (AI)", 
        "Điểm Cảm Xúc (10)", "Giá Gợi Ý Tối Ưu (đ)", "Lời Khuyên Mua Sắm", "Link Lazada"
    ]
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Ghi dữ liệu từng dòng
    for row_idx, p in enumerate(products, start=2):
        ai = p.ai_analysis or {}
        pros_str = " \n• ".join(ai.get("pros", [])) if isinstance(ai, dict) and ai.get("pros") else ""
        if pros_str:
            pros_str = f"• {pros_str}"
        cons_str = " \n• ".join(ai.get("cons", [])) if isinstance(ai, dict) and ai.get("cons") else ""
        if cons_str:
            cons_str = f"• {cons_str}"

        norm_name = ai.get("normalized_name", p.name) if isinstance(ai, dict) else p.name
        quality = ai.get("quality_summary", "") if isinstance(ai, dict) else ""
        sentiment = ai.get("sentiment_score", "") if isinstance(ai, dict) else ""
        p_opt = ai.get("recommended_price_optimal", 0.0) if isinstance(ai, dict) else 0.0
        verdict = ai.get("buying_verdict", "") if isinstance(ai, dict) else ""

        row_data = [
            p.id,
            p.name,
            norm_name,
            p.platform.upper(),
            p.current_price,
            p.original_price or p.current_price,
            f"{p.discount_percentage:.1f}%" if p.discount_percentage else "0%",
            p.historical_sold or 0,
            f"{p.rating_star or 0:.1f}★",
            p.shop_name or "Gian hàng Lazada",
            quality,
            pros_str,
            cons_str,
            sentiment,
            p_opt if p_opt > 0 else p.current_price,
            verdict,
            p.url
        ]
        ws.append(row_data)

        # Định dạng borders cho dòng dữ liệu
        for col_idx in range(1, len(headers) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.border = thin_border
            if col_idx in [1, 4, 7, 8, 9, 14]:
                c.alignment = align_center
            else:
                c.alignment = align_left

    # Tự động điều chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            first_line = val_str.split('\n')[0] if '\n' in val_str else val_str
            max_len = max(max_len, len(first_line))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"lazada_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/json", summary="Xuất danh sách sản phẩm ra tệp JSON chuẩn")
async def export_products_json(
    q: Optional[str] = Query(None),
    platform: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db)
):
    """Xuất danh sách sản phẩm kèm cấu trúc JSON AI hoàn chỉnh"""
    query = select(Product).order_by(desc(Product.id))
    if q and q.strip():
        search_term = f"%{q.strip()}%"
        query = query.where(or_(Product.name.ilike(search_term), Product.brand.ilike(search_term)))
    if platform and platform.lower() != "all":
        query = query.where(Product.platform == platform.lower())

    result = await db.execute(query)
    products = result.scalars().all()

    data = [serialize_product(p) for p in products]
    json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
    filename = f"lazada_products_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    return Response(
        content=json_bytes,
        media_type="application/json; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@router.get("/{product_id}", response_model=ProductResponse, summary="Xem chi tiết sản phẩm và lịch sử giá")
async def get_product_detail(
    product_id: int,
    db: AsyncSession = Depends(get_db)
):
    stmt = select(Product).where(Product.id == product_id)
    res = await db.execute(stmt)
    product = res.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Không tìm thấy sản phẩm")

    # Lấy lịch sử biến động giá
    history_stmt = (
        select(PriceHistory)
        .where(PriceHistory.product_id == product_id)
        .order_by(asc(PriceHistory.recorded_at))
    )
    hist_res = await db.execute(history_stmt)
    history_items = hist_res.scalars().all()

    serialized = serialize_product(product)
    serialized.price_history = [
        PriceHistoryResponse(
            id=h.id,
            product_id=h.product_id,
            price=h.price,
            original_price=h.original_price,
            discount_percentage=h.discount_percentage,
            scraped_at=h.recorded_at
        )
        for h in history_items
    ]
    return serialized


@router.get("/{product_id}/price-history", summary="Lấy dữ liệu chuỗi thời gian biến động giá để vẽ biểu đồ")
async def get_price_history_timeline(
    product_id: int,
    db: AsyncSession = Depends(get_db)
):
    stmt = (
        select(PriceHistory)
        .where(PriceHistory.product_id == product_id)
        .order_by(asc(PriceHistory.recorded_at))
    )
    result = await db.execute(stmt)
    history = result.scalars().all()

    if not history:
        # Nếu chưa có history riêng, lấy giá hiện tại của sản phẩm làm 1 điểm
        prod = (await db.execute(select(Product).where(Product.id == product_id))).scalar_one_or_none()
        if prod:
            ts = prod.updated_at.isoformat() if prod.updated_at else (prod.created_at.isoformat() if prod.created_at else "")
            return [{
                "timestamp": ts,
                "price": prod.current_price,
                "original_price": prod.original_price or prod.current_price,
                "discount_percentage": prod.discount_percentage or 0
            }]
        return []

    return [
        {
            "timestamp": h.recorded_at.isoformat() if h.recorded_at else "",
            "price": h.price,
            "original_price": h.original_price or h.price,
            "discount_percentage": h.discount_percentage or 0
        }
        for h in history
    ]
